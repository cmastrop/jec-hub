import json
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import unicodedata
import re

SUPABASE_URL = "https://dlpopmazvupiukpxurmf.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRscG9wbWF6dnVwaXVrcHh1cm1mIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MTA2MjM1NywiZXhwIjoyMDg2NjM4MzU3fQ.1br-NU_lr9pr2htfgh43OAww6bInIaKSEFtYC9MYk7E"

def fetch_all_songs():
    all_songs = []
    offset = 0
    limit = 1000
    while True:
        url = f"{SUPABASE_URL}/rest/v1/songs?select=id,title,artist,original_key,status,source_type,created_at&order=title.asc&offset={offset}&limit={limit}"
        req = urllib.request.Request(url, headers={
            "apikey": SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
        })
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
        if not data:
            break
        all_songs.extend(data)
        print(f"  Fetched {len(all_songs)} songs...")
        if len(data) < limit:
            break
        offset += limit
    return all_songs

def normalize(title):
    title = (title or "").lower()
    title = unicodedata.normalize("NFD", title)
    title = re.sub(r'[\u0300-\u036f]', '', title)
    title = re.sub(r'[^a-z0-9\s]', '', title)
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def main():
    print("Fetching songs from Supabase...")
    songs = fetch_all_songs()
    print(f"Total: {len(songs)} songs\n")

    # Detect duplicates (same normalized title, different keys)
    title_groups = {}
    for s in songs:
        norm = normalize(s["title"])
        if norm not in title_groups:
            title_groups[norm] = []
        title_groups[norm].append(s)

    dup_titles = {norm for norm, group in title_groups.items() if len(group) > 1}

    # Create workbook
    wb = openpyxl.Workbook()

    # --- Sheet 1: All Songs ---
    ws = wb.active
    ws.title = "Todas las Canciones"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    dup_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber-100
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    headers = ["#", "Titulo", "Artista", "Tonalidad", "Estado", "Origen", "Fecha", "Duplicado?"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    source_labels = {
        "manual": "Manual",
        "import_image": "Imagen",
        "import_pdf": "PDF",
        "import_dropbox": "Dropbox",
    }
    status_labels = {
        "draft": "Borrador",
        "published": "Publicada",
        "archived": "Archivada",
    }

    for i, song in enumerate(songs, 1):
        norm = normalize(song["title"])
        is_dup = norm in dup_titles
        keys_in_group = [s["original_key"] for s in title_groups[norm]]

        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=song["title"]).border = thin_border
        ws.cell(row=row, column=3, value=song.get("artist") or "").border = thin_border

        key_cell = ws.cell(row=row, column=4, value=song.get("original_key") or "")
        key_cell.border = thin_border
        key_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=5, value=status_labels.get(song.get("status", ""), song.get("status", ""))).border = thin_border
        ws.cell(row=row, column=6, value=source_labels.get(song.get("source_type", ""), song.get("source_type", ""))).border = thin_border

        date_str = (song.get("created_at") or "")[:10]
        ws.cell(row=row, column=7, value=date_str).border = thin_border

        if is_dup:
            unique_keys = sorted(set(keys_in_group))
            dup_text = f"Si ({len(title_groups[norm])} versiones: {', '.join(unique_keys)})"
            dup_cell = ws.cell(row=row, column=8, value=dup_text)
            dup_cell.border = thin_border
            dup_cell.fill = dup_fill
        else:
            ws.cell(row=row, column=8, value="No").border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:H{len(songs) + 1}"

    # --- Sheet 2: Duplicates Only ---
    ws2 = wb.create_sheet("Duplicados (diff tonalidad)")
    headers2 = ["Titulo (normalizado)", "Versiones", "Tonalidades"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    dup_row = 2
    for norm in sorted(dup_titles):
        group = title_groups[norm]
        keys = sorted(set(s["original_key"] for s in group))
        display_title = group[0]["title"]
        ws2.cell(row=dup_row, column=1, value=display_title).border = thin_border
        ws2.cell(row=dup_row, column=2, value=len(group)).border = thin_border
        ws2.cell(row=dup_row, column=3, value=", ".join(keys)).border = thin_border
        dup_row += 1

    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 25
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:C{dup_row - 1}"

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet("Resumen")
    summary_data = [
        ("Total de canciones", len(songs)),
        ("Canciones unicas (por titulo)", len(title_groups)),
        ("Canciones con multiples tonalidades", len(dup_titles)),
        ("Canciones unica tonalidad", len(title_groups) - len(dup_titles)),
        ("", ""),
        ("Por tonalidad", ""),
    ]

    key_counts = {}
    for s in songs:
        k = s.get("original_key") or "?"
        key_counts[k] = key_counts.get(k, 0) + 1
    for k in sorted(key_counts, key=lambda x: -key_counts[x]):
        summary_data.append((f"  {k}", key_counts[k]))

    for i, (label, val) in enumerate(summary_data, 1):
        cell_a = ws3.cell(row=i, column=1, value=label)
        cell_b = ws3.cell(row=i, column=2, value=val)
        if i == 1 or label == "Por tonalidad":
            cell_a.font = Font(bold=True, size=12)
        cell_a.border = thin_border
        cell_b.border = thin_border

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 15

    # Save
    output = "C:/Users/CMastropasqua/Downloads/JEC_HUB_Canciones.xlsx"
    wb.save(output)
    print(f"Excel saved to: {output}")
    print(f"  Sheet 1: {len(songs)} canciones")
    print(f"  Sheet 2: {len(dup_titles)} grupos con multiples tonalidades")
    print(f"  Sheet 3: Resumen general")

main()
